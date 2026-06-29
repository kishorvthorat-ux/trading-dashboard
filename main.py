import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import CellIsRule

# ===============================
# LOAD DATA
# ===============================
folder_path = "./data"

csv_files = [f for f in os.listdir(folder_path) if f.endswith(".csv")]
latest_file = max(csv_files, key=lambda x: os.path.getmtime(os.path.join(folder_path, x)))

df = pd.read_csv(os.path.join(folder_path, latest_file))
df['Date and time'] = pd.to_datetime(df['Date and time'])

# ===============================
# TRANSFORM
# ===============================
df = df.sort_values(['Trade number', 'Date and time'])

entries = df[df['Type'].str.contains("Entry")].copy()
exits = df[df['Type'].str.contains("Exit")].copy()

entries = entries.rename(columns={
    'Type': 'Entry Type',
    'Date and time': 'Entry Date',
    'Signal': 'Entry Signal',
    'Price USD': 'Entry Price'
})

exits = exits.rename(columns={
    'Type': 'Exit Type',
    'Date and time': 'Exit Date',
    'Signal': 'Exit Signal',
    'Price USD': 'Exit Price'
})

calc = pd.merge(entries, exits, on='Trade number')

# ===============================
# CALCULATIONS
# ===============================
calc['Points'] = calc['Exit Price'] - calc['Entry Price']
calc['Gross PnL'] = calc['Points'] * 30
calc['Turnover'] = calc['Entry Price'] * 30
calc['Brokerage'] = calc['Turnover'] * 0.0005
calc['Net PnL'] = calc['Gross PnL'] - calc['Brokerage']

calc['Equity'] = 100000 + calc['Net PnL'].cumsum()
calc['Return %'] = (calc['Points'] / calc['Entry Price']) * 100

# ===============================
# STATS
# ===============================
wins = (calc['Net PnL'] > 0).sum()
losses = (calc['Net PnL'] < 0).sum()

stats_df = pd.DataFrame([{
    "Total": len(calc),
    "Wins": wins,
    "Losses": losses,
    "Strike Rate": wins / len(calc),
    "Net PnL": calc['Net PnL'].sum()
}])

# ===============================
# SAVE TO EXCEL
# ===============================
file_name = "trading_output.xlsx"

with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
    calc.to_excel(writer, sheet_name="Calculated", index=False)
    stats_df.to_excel(writer, sheet_name="Stats", index=False)

# ===============================
# OPENPYXL FORMATTING
# ===============================
wb = load_workbook(file_name)

sheet = wb["Calculated"]

# Header style
header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

for cell in sheetcell.font = Font(bold=True)
    cell.fill = header_fill

# Auto width
for col in sheet.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    sheet.column_dimensions[col[0].column_letter].width = max_length + 2

# Conditional formatting for Net PnL
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

sheet.conditional_formatting.add(
    "F2:F1000",  # adjust column index if needed
    CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill)
)

sheet.conditional_formatting.add(
    "F2:F1000",
    CellIsRule(operator='lessThan', formula=['0'], fill=red_fill)
)

wb.save(file_name)

print("✅ Excel file with formatting created!")
